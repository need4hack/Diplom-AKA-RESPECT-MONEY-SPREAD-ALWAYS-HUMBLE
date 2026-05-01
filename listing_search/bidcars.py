from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import undetected_chromedriver as uc
import time
import xlsxwriter
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
#второй способ
import json
#третий способ
import cloudscraper
import asyncio
from playwright.async_api import async_playwright
from curl_cffi import requests
from curl_cffi.requests import Session
from playwright.sync_api import sync_playwright
import os
from openpyxl import Workbook, load_workbook

file_path = "c:/Users/stasp/OneDrive/Рабочий стол/парсер/Bid_cars_ТЕСТГОВНА31.01.xlsx"

headers1 = {
    "Authority": "bid.cars",
    "Method": "GET",
    "Path": "/app/search/request?search-type=filters&status=All&type=Automobile&make=All&model=All&year-from=2010&year-to=2010&auction-type=All",
    "Scheme": "https",
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br, zstd",
    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36",
    "Cookie": "_ga=GA1.1.1245070991.1763316192; _fbp=fb.1.1763316192170.936200483744354021; new_layout=jNWN5i6hrhxI; __cflb=02DiuFJ5hrNnjSxNqkRoDxfMegRhD1DPJtm18Jbq4T8Hi; bidcars_session=0dt43qt6osk3vns94q6sncdcjd; cf_clearance=iEkyVPaoZJXacN_tXvMZs.wrUbh1BuRRuz8n4di0lnU-1769547105-1.2.1.1-7zrWTuK5XelVSdK1ezhsl5WX4fc6MHaWy6._k4cMNXO.WqVlqtsnANx8nBqpgqlh2KBLU8jUenc.zFs7GFG.seX39HMLtpgUtha58mpsp2P4bzyoxoKoGjUIhur4zkDM1FCqhNABLiHr0Nwy6is2ORlZL_MCK0IROdqan8lUkAuu20VfwMT1pn77wXOPu8gN8BFjMQVxC7oBmkmaEjQ8vPSIBkAYAwwQYHZi3rAqlD73kEfSPzjjyBfXeY34NMWW; __cf_bm=uDJ8aksRI79e58BygylDcbP89YcWDMrEvCW4w_yWTv8-1769547705-1.0.1.1-sJa3qqlKnAzVU2yyUzR90SoCSE0dPd_EGOcBDe_jPCljLQr5MrMvE7rRVRUGE79Vtve3t2gdTe62JIqKr3a30l3tA25.a07n6kmanpvBKS8; _ga_TM0LLNB30Z=GS2.1.s1769545784$o10$g1$t1769548010$j60$l0$h0",
    #"Referer": "https://bid.cars/en/search/results?search-type=filters&status=All&type=Automobile&make=All&model=All&year-from=2023&year-to=2026&auction-type=All",
    "Referer": "https://bid.cars/en/search/results?search-type=filters&status=All&type=Automobile&make=All&model=All&year-from=2010&year-to=2010&auction-type=All",
    # "Origin": "https://bid.cars",
    "Sec-Fetch-Site": "same-origin",
}
# Если файла нет — создаём с заголовками
if not os.path.exists(file_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Lots"
    worksheet.append(["URL", "Lot-number", "VIN", "Name", "Engine info", "Key Info", "Drive Type", "Fuel Type", "Transmission"])
    workbook.save(file_path)

# Загружаем существующий файл
workbook = load_workbook(file_path)
worksheet = workbook.active
row = worksheet.max_row + 1


for m in range(1999, 2000):  
    url_years_base = f'year-from={m}&year-to={m}'
    for n in range(1, 51):
        # url1 = f'https://bid.cars/app/search/request?search-type=filters&status=All&type=Automobile&make=All&model=All&{url_years_base}&auction-type=All&page={n}'
        url1 = f'https://bid.cars/app/search/archived/request?search-type=filters&status=All&type=Automobile&make=All&model=All&{url_years_base}&auction-type=All&page={n}'
        # session = Session(
        #     #browser="chrome",
        #     impersonate="chrome110",
        #     verify=False)
        response = requests.get(url1, headers=headers1, impersonate="chrome110")
        #response = session.get(url1)
        if response.status_code != 200:
            print(f"Ошибка {response.status_code} при загрузке {url1}")
            continue

        data = response.json()
        for item in data.get('data', []):
            try:
                lot_number = item['lot']
                name_long = item['name_long']
                engine_info = item['specs']['engine_rendered']
                drive_type = item['specs']['drive_type']
                fuel_type = item['specs']['fuel_type']
                transmission = item['specs']['transmission']
                key_info = item['specs']['key_info']
                vin = item['vin']
                url2 = f"https://bid.cars/en/lot/{lot_number}"

                worksheet.append([url2, lot_number, vin, name_long, engine_info, key_info, drive_type, fuel_type, transmission])
                print(f"✓ {lot_number} записан")
            except Exception as e:
                print(f"Ошибка обработки данных: {e}")

# Сохраняем файл после всего
workbook.save(file_path)
print("✅ Файл успешно обновлён.")
